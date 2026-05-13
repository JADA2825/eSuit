"""
cuadro_cargas.py — Generación del Cuadro de Cargas profesional.

Distribuye los circuitos del proyecto entre las fases R, S y T y calcula
balanceo, % desbalance y demanda total. También genera el Excel con fórmulas
(no valores) para edición manual.
"""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════
# ALGORITMO DE BALANCEO DE FASES
# ════════════════════════════════════════════════════
def asignar_fases(circuitos: list) -> list:
    """Asigna cada circuito a una o más fases (R, S, T) usando un algoritmo
    greedy: en cada paso asigna el circuito mas grande a la(s) fase(s) menos
    cargada(s).

    Retorna la lista de circuitos con los campos:
        - 'fases'  : lista con 1, 2 o 3 fases (subset de ['R','S','T'])
        - 'P_R'    : potencia en fase R (W)
        - 'P_S'    : potencia en fase S (W)
        - 'P_T'    : potencia en fase T (W)
    """
    carga_acum = {"R": 0.0, "S": 0.0, "T": 0.0}
    # Procesar de mayor a menor potencia para balancear mejor
    indices_orden = sorted(
        range(len(circuitos)),
        key=lambda i: circuitos[i].get("potencia", 0) or 0,
        reverse=True,
    )
    out = [dict(c) for c in circuitos]

    for i in indices_orden:
        c = out[i]
        P = float(c.get("potencia", 0) or 0)
        config_code = c.get("config_code", "mono2h")

        if config_code == "trifasico":
            # Las 3 fases reciben P/3 cada una
            fases = ["R", "S", "T"]
            cuota = P / 3.0
            P_R = P_S = P_T = cuota
            carga_acum["R"] += cuota
            carga_acum["S"] += cuota
            carga_acum["T"] += cuota
        elif config_code == "mono3h":
            # 2 fases (mono 3 hilos): P/2 a cada una de las 2 fases con menor carga
            fases_ordenadas = sorted(carga_acum.keys(), key=lambda f: carga_acum[f])
            fases = fases_ordenadas[:2]
            cuota = P / 2.0
            P_R = cuota if "R" in fases else 0
            P_S = cuota if "S" in fases else 0
            P_T = cuota if "T" in fases else 0
            for f in fases:
                carga_acum[f] += cuota
        else:  # mono2h
            # 1 fase (fase + neutro): asignar a la fase menos cargada
            fase = min(carga_acum.keys(), key=lambda f: carga_acum[f])
            fases = [fase]
            P_R = P if fase == "R" else 0
            P_S = P if fase == "S" else 0
            P_T = P if fase == "T" else 0
            carga_acum[fase] += P

        c["fases"] = fases
        c["P_R"] = round(P_R, 1)
        c["P_S"] = round(P_S, 1)
        c["P_T"] = round(P_T, 1)

    return out


def resumen_balanceo(circuitos_con_fases: list) -> dict:
    """Devuelve totales por fase y % desbalance."""
    P_R = sum(c.get("P_R", 0) for c in circuitos_con_fases)
    P_S = sum(c.get("P_S", 0) for c in circuitos_con_fases)
    P_T = sum(c.get("P_T", 0) for c in circuitos_con_fases)
    P_total = P_R + P_S + P_T
    P_prom = P_total / 3.0 if P_total > 0 else 0
    P_max  = max(P_R, P_S, P_T)
    P_min  = min(P_R, P_S, P_T)

    # Desbalance NEMA / IEEE: ((P_max - P_prom) / P_prom) × 100
    if P_prom > 0:
        desbalance = max(P_max - P_prom, P_prom - P_min) / P_prom * 100
    else:
        desbalance = 0.0

    # Calidad del balanceo
    if desbalance < 5:
        calidad = "Excelente"
        calidad_color = "verde"
    elif desbalance < 10:
        calidad = "Aceptable"
        calidad_color = "verde"
    elif desbalance < 20:
        calidad = "Regular"
        calidad_color = "amarillo"
    else:
        calidad = "Deficiente"
        calidad_color = "rojo"

    return {
        "P_R": P_R, "P_S": P_S, "P_T": P_T,
        "P_total": P_total, "P_prom": P_prom,
        "P_max_fase": P_max, "P_min_fase": P_min,
        "desbalance_pct": round(desbalance, 2),
        "calidad": calidad,
        "calidad_color": calidad_color,
    }


# ════════════════════════════════════════════════════
# EXCEL CON FÓRMULAS (no valores fijos)
# ════════════════════════════════════════════════════

# Estilos
FILL_HEADER = PatternFill("solid", fgColor="1F3864")
FILL_SUB    = PatternFill("solid", fgColor="2D4A7A")
FILL_PHASE_R = PatternFill("solid", fgColor="FFE4E1")  # rosa claro
FILL_PHASE_S = PatternFill("solid", fgColor="E4F1FE")  # azul claro
FILL_PHASE_T = PatternFill("solid", fgColor="FFF4E4")  # naranja claro
FILL_ALT    = PatternFill("solid", fgColor="F5F5F7")
FILL_OK     = PatternFill("solid", fgColor="DCFCE7")
FILL_WARN   = PatternFill("solid", fgColor="FEF3C7")
FILL_ERR    = PatternFill("solid", fgColor="FEE2E2")
FILL_TOTAL  = PatternFill("solid", fgColor="DBEAFE")

FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_SUB    = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
FONT_BODY   = Font(name="Calibri", size=10)
FONT_BOLD   = Font(name="Calibri", size=10, bold=True)
FONT_TOTAL  = Font(name="Calibri", size=11, bold=True, color="1F3864")
FONT_TITLE  = Font(name="Calibri", size=14, bold=True, color="1F3864")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")

THIN  = Side(border_style="thin", color="DCDCDC")
THICK = Side(border_style="medium", color="1F3864")
BORDER_CELL  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_THICK = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generar_excel_cuadro_cargas(circuitos: list, datos_proyecto: dict) -> bytes:
    """Genera un Excel profesional con:
       - Hoja "Cuadro de Cargas": tabla principal con fórmulas vivas
       - Hoja "Balanceo": resumen por fase con fórmulas SUMIF
       - Hoja "Circuitos detalle": datos crudos para auditoría
    """
    circs = asignar_fases(circuitos)
    bal = resumen_balanceo(circs)

    wb = Workbook()

    # ═══════════════════════════════════════════════════
    # HOJA 1 — CUADRO DE CARGAS (con fórmulas)
    # ═══════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Cuadro de Cargas"

    # ── Encabezado del proyecto (filas 1-5) ─────────
    ws["A1"] = f"CUADRO DE CARGAS — {datos_proyecto.get('proyecto','—')}"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:P1")
    ws["A1"].alignment = ALIGN_CENTER

    ws["A2"] = "Cliente:"
    ws["B2"] = datos_proyecto.get("cliente", "—")
    ws["E2"] = "Ubicación:"
    ws["F2"] = datos_proyecto.get("ubicacion", "—")
    ws["I2"] = "Sistema:"
    ws["J2"] = datos_proyecto.get("sistema", "—")
    ws["M2"] = "Norma:"
    ws["N2"] = datos_proyecto.get("norma", "NOM-001-SEDE-2012")

    ws["A3"] = "Responsable:"
    ws["B3"] = datos_proyecto.get("responsable", "—")
    ws["E3"] = "Cédula:"
    ws["F3"] = datos_proyecto.get("cedula", "—")
    ws["I3"] = "Fecha:"
    ws["J3"] = str(datos_proyecto.get("fecha", date.today()))
    ws["M3"] = "Temp. amb.:"
    ws["N3"] = f"{datos_proyecto.get('temp_ambiente', 35)} °C"

    for r in (2, 3):
        for col in ("A", "E", "I", "M"):
            ws[f"{col}{r}"].font = FONT_BOLD

    # ── Headers de la tabla (fila 5-6) ──────────────
    HEADERS = [
        "No.", "Descripción / Circuito",
        "P (W)", "V (V)", "FP",
        "Sist.", "L (m)",
        "Fases",
        "P_R (W)", "P_S (W)", "P_T (W)",
        "I (A)", "I.dis (A)",
        "Conductor", "Tierra AWG", "Protección",
    ]
    HEADER_ROW = 5
    for col_i, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_CELL
    ws.row_dimensions[HEADER_ROW].height = 30

    # ── Filas de circuitos (con FÓRMULAS) ───────────
    start_row = HEADER_ROW + 1
    for i, c in enumerate(circs):
        r = start_row + i
        fases_str = "+".join(c.get("fases", []))
        ws.cell(row=r, column=1, value=c.get("id", i + 1)).font = FONT_BODY
        ws.cell(row=r, column=2, value=c.get("nombre", f"Circuito {i+1}")).font = FONT_BODY
        # Datos de entrada (editables a mano)
        ws.cell(row=r, column=3, value=c.get("potencia", 0))                    # P
        ws.cell(row=r, column=4, value=c.get("voltaje", 0))                      # V
        ws.cell(row=r, column=5, value=c.get("fp", 0.9))                         # FP
        ws.cell(row=r, column=6, value=(c.get("config_code", "mono2h")
                                          .replace("mono", "Mono ")
                                          .replace("trifasico", "Trifásico")))   # Sist
        ws.cell(row=r, column=7, value=c.get("longitud", 0))                     # L
        ws.cell(row=r, column=8, value=fases_str).font = FONT_BOLD               # Fases
        ws.cell(row=r, column=9,  value=c.get("P_R", 0))
        ws.cell(row=r, column=10, value=c.get("P_S", 0))
        ws.cell(row=r, column=11, value=c.get("P_T", 0))
        # Corriente: FÓRMULA que depende de P, V, FP, Sist
        # Si "Trifásico" en F: =C/(SQRT(3)*D*E); si "Mono 3H" o "Bifásico": =C/(D*E); si "Mono 2H": =C/(D*E)
        config_code = c.get("config_code", "mono2h")
        if config_code == "trifasico":
            ws.cell(row=r, column=12, value=f"=IFERROR(C{r}/(SQRT(3)*D{r}*E{r}),0)")
        else:
            ws.cell(row=r, column=12, value=f"=IFERROR(C{r}/(D{r}*E{r}),0)")
        # Corriente de diseño: =L*1.25 si carga continua, sino =L*1.0
        fd = c.get("factor_demanda", 1.25)
        ws.cell(row=r, column=13, value=f"=L{r}*{fd}")
        # Conductor / Tierra / Protección — texto descriptivo
        cond_str = f"{c.get('cf',1)}× {c.get('conductor','—')} {c.get('material','')}"
        ws.cell(row=r, column=14, value=cond_str).font = FONT_BODY
        ws.cell(row=r, column=15, value=f"{c.get('tierra_calibre','—')} {c.get('tierra_material','Cu')}")
        ws.cell(row=r, column=16, value=c.get("proteccion_fmt",
                                                f"{c.get('proteccion_A','—')} A"))

        # Estilos por fila
        fila_fill = FILL_ALT if (i % 2 == 1) else None
        for col_i in range(1, 17):
            cell = ws.cell(row=r, column=col_i)
            cell.border = BORDER_CELL
            cell.alignment = ALIGN_CENTER if col_i != 2 else ALIGN_LEFT
            if cell.font.name is None or not cell.font.name:
                cell.font = FONT_BODY
            if fila_fill:
                cell.fill = fila_fill
        # Resaltar columnas de fase
        ws.cell(row=r, column=9).fill = FILL_PHASE_R
        ws.cell(row=r, column=10).fill = FILL_PHASE_S
        ws.cell(row=r, column=11).fill = FILL_PHASE_T
        # Formato numérico
        for col_i in (3, 9, 10, 11):
            ws.cell(row=r, column=col_i).number_format = "#,##0"
        for col_i in (12, 13):
            ws.cell(row=r, column=col_i).number_format = "0.00"
        ws.cell(row=r, column=5).number_format = "0.00"
        ws.cell(row=r, column=7).number_format = "0"

    # ── Fila de TOTALES (con fórmulas SUM) ──────────
    total_row = start_row + len(circs) + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = FONT_TOTAL
    ws.cell(row=total_row, column=2, value=f"{len(circs)} circuito(s)").font = FONT_TOTAL
    ws.cell(row=total_row, column=3,
            value=f"=SUM(C{start_row}:C{start_row+len(circs)-1})")
    for col_i, col_letter in [(9, "I"), (10, "J"), (11, "K"),
                                (12, "L"), (13, "M")]:
        ws.cell(row=total_row, column=col_i,
                value=f"=SUM({col_letter}{start_row}:{col_letter}{start_row+len(circs)-1})")
    for col_i in range(1, 17):
        cell = ws.cell(row=total_row, column=col_i)
        cell.font = FONT_TOTAL
        cell.fill = FILL_TOTAL
        cell.border = BORDER_THICK
        cell.alignment = ALIGN_CENTER if col_i != 2 else ALIGN_LEFT
        if col_i in (3, 9, 10, 11):
            cell.number_format = "#,##0"
        elif col_i in (12, 13):
            cell.number_format = "0.00"

    # ── Sección de BALANCEO (debajo del cuadro) ─────
    bal_row = total_row + 3
    ws.cell(row=bal_row, column=1, value="BALANCEO DE FASES").font = FONT_TITLE
    ws.merge_cells(start_row=bal_row, start_column=1, end_row=bal_row, end_column=11)
    ws.cell(row=bal_row, column=1).alignment = ALIGN_LEFT

    bal_headers_row = bal_row + 1
    ws.cell(row=bal_headers_row, column=1, value="Fase").font = FONT_HEADER
    ws.cell(row=bal_headers_row, column=2, value="P (W)").font = FONT_HEADER
    ws.cell(row=bal_headers_row, column=3, value="% del total").font = FONT_HEADER
    ws.cell(row=bal_headers_row, column=4, value="vs promedio").font = FONT_HEADER
    for c in range(1, 5):
        ws.cell(row=bal_headers_row, column=c).fill = FILL_HEADER
        ws.cell(row=bal_headers_row, column=c).alignment = ALIGN_CENTER
        ws.cell(row=bal_headers_row, column=c).border = BORDER_CELL

    P_R_cell = f"I{total_row}"
    P_S_cell = f"J{total_row}"
    P_T_cell = f"K{total_row}"
    P_total_formula = f"({P_R_cell}+{P_S_cell}+{P_T_cell})"
    P_prom_formula  = f"(({P_R_cell}+{P_S_cell}+{P_T_cell})/3)"

    fase_rows = [
        ("R", P_R_cell, FILL_PHASE_R),
        ("S", P_S_cell, FILL_PHASE_S),
        ("T", P_T_cell, FILL_PHASE_T),
    ]
    for fi, (fase, cell_ref, fill) in enumerate(fase_rows):
        rr = bal_headers_row + 1 + fi
        ws.cell(row=rr, column=1, value=fase).font = FONT_BOLD
        ws.cell(row=rr, column=2, value=f"={cell_ref}").number_format = "#,##0"
        ws.cell(row=rr, column=3,
                value=f"=IFERROR({cell_ref}/{P_total_formula},0)").number_format = "0.00%"
        ws.cell(row=rr, column=4,
                value=f"=IFERROR(({cell_ref}-{P_prom_formula})/{P_prom_formula},0)"
                ).number_format = "0.00%"
        for c in range(1, 5):
            ws.cell(row=rr, column=c).alignment = ALIGN_CENTER
            ws.cell(row=rr, column=c).border = BORDER_CELL
            ws.cell(row=rr, column=c).fill = fill

    # Total y % desbalance
    rr_total = bal_headers_row + 4
    ws.cell(row=rr_total, column=1, value="TOTAL").font = FONT_TOTAL
    ws.cell(row=rr_total, column=2,
            value=f"={P_total_formula}").number_format = "#,##0"
    ws.cell(row=rr_total, column=3, value="100%").alignment = ALIGN_CENTER
    ws.cell(row=rr_total, column=4, value="—").alignment = ALIGN_CENTER
    for c in range(1, 5):
        ws.cell(row=rr_total, column=c).font = FONT_TOTAL
        ws.cell(row=rr_total, column=c).fill = FILL_TOTAL
        ws.cell(row=rr_total, column=c).border = BORDER_THICK
        ws.cell(row=rr_total, column=c).alignment = ALIGN_CENTER

    # % desbalance (NEMA)
    rr_desb = rr_total + 2
    ws.cell(row=rr_desb, column=1, value="% Desbalance:").font = FONT_BOLD
    # = MAX(ABS(P-prom))/prom
    formula_desb = (
        f"=IFERROR(MAX(ABS({P_R_cell}-{P_prom_formula}),"
        f"ABS({P_S_cell}-{P_prom_formula}),"
        f"ABS({P_T_cell}-{P_prom_formula}))/{P_prom_formula},0)"
    )
    ws.cell(row=rr_desb, column=2, value=formula_desb).number_format = "0.00%"
    ws.cell(row=rr_desb, column=2).font = FONT_TOTAL

    # Calidad del balanceo (texto basado en %)
    ws.cell(row=rr_desb, column=3, value="Criterio:").font = FONT_BOLD
    ws.cell(row=rr_desb, column=4,
            value=f'=IF(B{rr_desb}<5%,"Excelente",'
                  f'IF(B{rr_desb}<10%,"Aceptable",'
                  f'IF(B{rr_desb}<20%,"Regular","Deficiente")))').font = FONT_BOLD

    # ── Demanda total (kVA y I) ────────────────────
    rr_dem = rr_desb + 3
    ws.cell(row=rr_dem, column=1, value="DEMANDA TOTAL").font = FONT_TITLE
    ws.merge_cells(start_row=rr_dem, start_column=1, end_row=rr_dem, end_column=4)

    rr_dem2 = rr_dem + 1
    ws.cell(row=rr_dem2, column=1, value="Carga total instalada:").font = FONT_BOLD
    ws.cell(row=rr_dem2, column=2,
            value=f"={P_total_formula}/1000").number_format = '#,##0.00" kW"'

    rr_dem3 = rr_dem2 + 1
    ws.cell(row=rr_dem3, column=1, value="Carga aparente (S):").font = FONT_BOLD
    # S = P / FP_promedio  — usar 0.9 como aproximación; el usuario puede editar
    ws.cell(row=rr_dem3, column=2, value=f"={P_total_formula}/1000/0.9"
            ).number_format = '#,##0.00" kVA"'

    # Anchos de columna
    _set_col_widths(ws, [6, 30, 11, 9, 7, 11, 8, 8, 11, 11, 11, 10, 11, 18, 13, 14])

    # Congelar paneles (fila de header + 2 columnas)
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=3)

    # ═══════════════════════════════════════════════════
    # HOJA 2 — BALANCEO (resumen visual con SUMIF)
    # ═══════════════════════════════════════════════════
    ws2 = wb.create_sheet("Balanceo")
    ws2["A1"] = "RESUMEN DE BALANCEO DE FASES"
    ws2["A1"].font = FONT_TITLE
    ws2.merge_cells("A1:F1")
    ws2["A1"].alignment = ALIGN_CENTER

    ws2["A3"] = "Fase"
    ws2["B3"] = "Potencia (W)"
    ws2["C3"] = "% del total"
    ws2["D3"] = "Desviación"
    for c in ("A3", "B3", "C3", "D3"):
        ws2[c].font = FONT_HEADER
        ws2[c].fill = FILL_HEADER
        ws2[c].alignment = ALIGN_CENTER
        ws2[c].border = BORDER_CELL

    # Referencias a la hoja 1 (fórmulas)
    sheet1 = "'Cuadro de Cargas'"
    for fi, (fase, col_letter, fill) in enumerate([
        ("R", "I", FILL_PHASE_R),
        ("S", "J", FILL_PHASE_S),
        ("T", "K", FILL_PHASE_T),
    ]):
        rr = 4 + fi
        ws2.cell(row=rr, column=1, value=fase).font = FONT_BOLD
        ws2.cell(row=rr, column=2,
                 value=f"=SUM({sheet1}!{col_letter}{start_row}:{col_letter}{start_row+len(circs)-1})"
                 ).number_format = "#,##0"
        total_ref = f"(B4+B5+B6)"
        ws2.cell(row=rr, column=3,
                 value=f"=IFERROR(B{rr}/{total_ref},0)").number_format = "0.00%"
        ws2.cell(row=rr, column=4,
                 value=f"=IFERROR((B{rr}-{total_ref}/3)/({total_ref}/3),0)"
                 ).number_format = "0.00%"
        for cc in range(1, 5):
            ws2.cell(row=rr, column=cc).alignment = ALIGN_CENTER
            ws2.cell(row=rr, column=cc).border = BORDER_CELL
            ws2.cell(row=rr, column=cc).fill = fill

    ws2["A7"] = "TOTAL"
    ws2["B7"] = "=B4+B5+B6"
    ws2["B7"].number_format = "#,##0"
    ws2["C7"] = "100%"
    for c in ("A7", "B7", "C7", "D7"):
        ws2[c].font = FONT_TOTAL
        ws2[c].fill = FILL_TOTAL
        ws2[c].border = BORDER_THICK
        ws2[c].alignment = ALIGN_CENTER

    ws2["A9"] = "% Desbalance (NEMA):"
    ws2["B9"] = (
        "=IFERROR(MAX(ABS(B4-(B4+B5+B6)/3),"
        "ABS(B5-(B4+B5+B6)/3),"
        "ABS(B6-(B4+B5+B6)/3))/((B4+B5+B6)/3),0)"
    )
    ws2["B9"].number_format = "0.00%"
    ws2["B9"].font = FONT_TOTAL
    ws2["A9"].font = FONT_BOLD

    ws2["A10"] = "Criterio:"
    ws2["B10"] = (
        '=IF(B9<5%,"Excelente",IF(B9<10%,"Aceptable",'
        'IF(B9<20%,"Regular","Deficiente")))'
    )
    ws2["B10"].font = FONT_TOTAL
    ws2["A10"].font = FONT_BOLD

    ws2["A12"] = "Recomendaciones NOM-001:"
    ws2["A12"].font = FONT_BOLD
    ws2["A13"] = "• < 5%: Balanceo excelente. Ideal para alimentadores."
    ws2["A14"] = "• 5-10%: Aceptable para tableros derivados."
    ws2["A15"] = "• 10-20%: Reasignar circuitos para balancear."
    ws2["A16"] = "• > 20%: Acción obligatoria; afecta vida útil de equipos."

    _set_col_widths(ws2, [28, 16, 16, 16])

    # ═══════════════════════════════════════════════════
    # HOJA 3 — CIRCUITOS DETALLE (datos crudos)
    # ═══════════════════════════════════════════════════
    ws3 = wb.create_sheet("Detalle circuitos")
    headers3 = [
        "No.", "Circuito", "Sistema", "P (W)", "V (V)", "FP", "L (m)",
        "I real (A)", "I.dis (A)", "Material", "CF", "Conductor",
        "Amp. corr. (A)", "CdT (%)", "CdT máx (%)", "Estado",
        "OCPD", "OCPD comercial",
        "Tierra calibre", "Tierra material",
        "Canalización", "T° term.", "Aislamiento",
        "Tubería", "Relleno tubería (%)",
        "Fases", "P_R", "P_S", "P_T",
    ]
    for col_i, h in enumerate(headers3, start=1):
        cell = ws3.cell(row=1, column=col_i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_CELL
    ws3.row_dimensions[1].height = 28

    for i, c in enumerate(circs):
        r = 2 + i
        rec = (c.get("conduit_info") or {}).get("recomendada") or {}
        cdt_v = c.get("cdt", 0) or 0
        cdt_mx = c.get("cdt_max", 3.0)
        vals = [
            c.get("id", i+1),
            c.get("nombre", f"Cto {i+1}"),
            c.get("configuracion", "—"),
            c.get("potencia", 0),
            c.get("voltaje", 0),
            c.get("fp", 0.9),
            c.get("longitud", 0),
            round(c.get("corriente", 0) or 0, 2),
            round(c.get("corriente_diseño", 0) or 0, 2),
            c.get("material", "—"),
            c.get("cf", 1),
            c.get("conductor", "—"),
            round(c.get("ampacity_corr", 0) or 0, 1),
            round(cdt_v, 2),
            cdt_mx,
            "Cumple" if cdt_v <= cdt_mx else "Revisar",
            c.get("proteccion_A", "—"),
            c.get("proteccion_fmt", "—"),
            c.get("tierra_calibre", "—"),
            c.get("tierra_material", "Cu"),
            c.get("canalizacion_label", c.get("canalizacion", "—")),
            f"{c.get('temp_term','—')}°C",
            f"{c.get('temp_aislamiento',75)}°C",
            rec.get("tubo", "—"),
            rec.get("fill_pct", "—"),
            "+".join(c.get("fases", [])),
            c.get("P_R", 0),
            c.get("P_S", 0),
            c.get("P_T", 0),
        ]
        for col_i, v in enumerate(vals, start=1):
            cell = ws3.cell(row=r, column=col_i, value=v)
            cell.border = BORDER_CELL
            cell.alignment = ALIGN_CENTER if col_i != 2 else ALIGN_LEFT
            cell.font = FONT_BODY
            if i % 2 == 1:
                cell.fill = FILL_ALT
        # Resaltar Estado
        estado_cell = ws3.cell(row=r, column=16)
        estado_cell.fill = FILL_OK if cdt_v <= cdt_mx else FILL_ERR
        estado_cell.font = FONT_BOLD
    _set_col_widths(ws3, [5, 22, 16, 10, 8, 6, 7, 9, 9, 9, 5, 11, 11, 9, 9, 10,
                          7, 11, 10, 12, 18, 9, 11, 14, 13, 8, 9, 9, 9])
    ws3.freeze_panes = ws3["C2"]

    # Salida
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
